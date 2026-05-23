import os
import json
import csv
import io
import base64
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

app = Flask(__name__)
CORS(app)

client = Groq(api_key=os.environ.get("GROQ_API_KEY"))

EXTRACTION_PROMPT = """You are an expert invoice data extractor. Analyze this invoice image and extract ALL data.

Return ONLY a valid JSON object (no markdown, no explanation) with this exact structure:
{
  "invoice_number": "string or null",
  "invoice_date": "YYYY-MM-DD or null",
  "due_date": "YYYY-MM-DD or null",
  "vendor": {
    "name": "string",
    "address": "string or null",
    "email": "string or null",
    "phone": "string or null",
    "tax_id": "string or null"
  },
  "bill_to": {
    "name": "string or null",
    "address": "string or null",
    "email": "string or null"
  },
  "line_items": [
    {
      "description": "string",
      "quantity": number,
      "unit_price": number,
      "total": number,
      "tax_rate": number or null
    }
  ],
  "subtotal": number,
  "tax_amount": number or null,
  "discount": number or null,
  "total_amount": number,
  "currency": "USD",
  "payment_terms": "string or null",
  "notes": "string or null",
  "confidence": "high|medium|low"
}

Be precise with numbers. Extract exact values shown. Use null for missing fields."""


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "message": "AI Invoice Reader API is running"})


@app.route("/api/extract", methods=["POST"])
def extract_invoice():
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "No file selected"}), 400

        allowed_types = ["image/jpeg", "image/jpg", "image/png", "image/webp", "image/gif"]
        if file.content_type not in allowed_types:
            return jsonify({"error": f"Unsupported file type: {file.content_type}. Use JPG, PNG, or WebP."}), 400

        file_content = file.read()
        base64_image = base64.b64encode(file_content).decode("utf-8")
        media_type = file.content_type

        response = client.chat.completions.create(
            model="meta-llama/llama-4-scout-17b-16e-instruct",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{media_type};base64,{base64_image}"
                            }
                        },
                        {
                            "type": "text",
                            "text": EXTRACTION_PROMPT
                        }
                    ]
                }
            ],
            max_tokens=2000,
            temperature=0.1
        )

        raw_text = response.choices[0].message.content.strip()

        # Clean up response
        if raw_text.startswith("```json"):
            raw_text = raw_text[7:]
        if raw_text.startswith("```"):
            raw_text = raw_text[3:]
        if raw_text.endswith("```"):
            raw_text = raw_text[:-3]
        raw_text = raw_text.strip()

        invoice_data = json.loads(raw_text)
        return jsonify({"success": True, "data": invoice_data})

    except json.JSONDecodeError as e:
        return jsonify({"error": f"Failed to parse AI response: {str(e)}", "raw": raw_text}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export/csv", methods=["POST"])
def export_csv():
    try:
        data = request.json
        invoices = data.get("invoices", [])

        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow([
            "Invoice #", "Date", "Due Date", "Vendor", "Vendor Email",
            "Bill To", "Description", "Quantity", "Unit Price",
            "Line Total", "Tax Rate", "Subtotal", "Tax Amount",
            "Discount", "Total Amount", "Currency", "Payment Terms", "Notes"
        ])

        for inv in invoices:
            items = inv.get("line_items", [{}])
            for i, item in enumerate(items):
                writer.writerow([
                    inv.get("invoice_number", "") if i == 0 else "",
                    inv.get("invoice_date", "") if i == 0 else "",
                    inv.get("due_date", "") if i == 0 else "",
                    inv.get("vendor", {}).get("name", "") if i == 0 else "",
                    inv.get("vendor", {}).get("email", "") if i == 0 else "",
                    inv.get("bill_to", {}).get("name", "") if i == 0 else "",
                    item.get("description", ""),
                    item.get("quantity", ""),
                    item.get("unit_price", ""),
                    item.get("total", ""),
                    item.get("tax_rate", ""),
                    inv.get("subtotal", "") if i == 0 else "",
                    inv.get("tax_amount", "") if i == 0 else "",
                    inv.get("discount", "") if i == 0 else "",
                    inv.get("total_amount", "") if i == 0 else "",
                    inv.get("currency", "USD") if i == 0 else "",
                    inv.get("payment_terms", "") if i == 0 else "",
                    inv.get("notes", "") if i == 0 else "",
                ])

        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode()),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"invoices_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export/xlsx", methods=["POST"])
def export_xlsx():
    try:
        data = request.json
        invoices = data.get("invoices", [])

        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary ──
        ws_summary = wb.active
        ws_summary.title = "Invoice Summary"

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        accent_fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")

        summary_headers = [
            "Invoice #", "Date", "Due Date", "Vendor", "Bill To",
            "Subtotal", "Tax", "Discount", "Total", "Currency", "Confidence"
        ]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, inv in enumerate(invoices, 2):
            fill = accent_fill if row_idx % 2 == 0 else PatternFill()
            values = [
                inv.get("invoice_number", "N/A"),
                inv.get("invoice_date", ""),
                inv.get("due_date", ""),
                inv.get("vendor", {}).get("name", ""),
                inv.get("bill_to", {}).get("name", ""),
                inv.get("subtotal", 0),
                inv.get("tax_amount", 0),
                inv.get("discount", 0),
                inv.get("total_amount", 0),
                inv.get("currency", "USD"),
                inv.get("confidence", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws_summary.cell(row=row_idx, column=col, value=val)
                cell.fill = fill
                if col in [6, 7, 8, 9]:
                    cell.number_format = '#,##0.00'

        for col in ws_summary.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws_summary.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ── Sheet 2: Line Items ──
        ws_items = wb.create_sheet("Line Items")
        item_headers = [
            "Invoice #", "Vendor", "Description", "Quantity",
            "Unit Price", "Total", "Tax Rate %"
        ]
        for col, header in enumerate(item_headers, 1):
            cell = ws_items.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row_idx = 2
        for inv in invoices:
            for item in inv.get("line_items", []):
                values = [
                    inv.get("invoice_number", "N/A"),
                    inv.get("vendor", {}).get("name", ""),
                    item.get("description", ""),
                    item.get("quantity", 0),
                    item.get("unit_price", 0),
                    item.get("total", 0),
                    item.get("tax_rate", 0),
                ]
                for col, val in enumerate(values, 1):
                    cell = ws_items.cell(row=row_idx, column=col, value=val)
                    if col in [5, 6]:
                        cell.number_format = '#,##0.00'
                row_idx += 1

        for col in ws_items.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws_items.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"invoices_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export/json", methods=["POST"])
def export_json():
    try:
        data = request.json
        invoices = data.get("invoices", [])
        output = json.dumps({"exported_at": datetime.now().isoformat(), "invoices": invoices}, indent=2)
        return send_file(
            io.BytesIO(output.encode()),
            mimetype="application/json",
            as_attachment=True,
            download_name=f"invoices_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
